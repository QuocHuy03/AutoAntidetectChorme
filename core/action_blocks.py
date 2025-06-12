import json
import time, os
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
import random
import pandas as pd
from core.api_bridge import close_profile
import openpyxl
import re

def render(text, local_vars):
    if not isinstance(text, str):
        return text

    combined_vars = {}
    if "row_data" in local_vars:
        combined_vars.update(local_vars["row_data"])
    else:
        combined_vars.update(local_vars)

    for key, val in combined_vars.items():
        if pd.notna(val):
            text = text.replace(f"{{{{{key}}}}}", str(val))

    return text

# Hàm thay thế tất cả các biến động trong chuỗi
def replace_variables_in_string(text, variables):
    return re.sub(r'{{(.*?)}}', lambda match: str(variables.get(match.group(1), match.group(0))), text)

def execute_blocks_from_json(json_path, logger, driver_path, debugger_address, profile_input, provider, base_url, stop_flag,
                             excel_mode='manual', excel_path=None):

    with open(json_path, 'r', encoding='utf-8') as f:
        blocks = json.load(f)

    def execute_block(block, local_vars):
        nonlocal variables, loop_flags

        if not isinstance(block, dict):
            logger(f"⚠️ Bỏ qua block không hợp lệ: {block}")
            return
        

        # if excel
        action = block.get('action')
        if action == 'excel':
            excel_path = block.get('path')
            mode = block.get('mode', 'row')
            inner_blocks = block.get('do', [])

            if not isinstance(inner_blocks, list):
                logger(f"[EXCEL BLOCK] ❌ 'do' không phải danh sách: {inner_blocks}")
                return

            try:
                df = pd.read_excel(excel_path)

                if mode == 'profile':
                    current_name = profile_input.get("name", "").strip().lower()
                    matched = False

                    for idx, row in df.iterrows():
                        if stop_flag.is_set():
                            logger(f"[{current_name}] ⛔ Dừng theo yêu cầu trong Excel block.")
                            return

                        row_profile = str(row.get("PROFILE", "")).strip().lower()
                        if row_profile != current_name:
                            continue

                        matched = True
                        row_vars = row.to_dict()
                        row_vars["row_data"] = row_vars.copy()

                        for inner in inner_blocks:
                            if stop_flag.is_set():
                                logger(f"[{current_name}] ⛔ Dừng giữa block Excel.")
                                return
                            execute_block(inner, row_vars)
                        break  # Dừng sau khi chạy đúng dòng trùng

                    if not matched:
                        logger(f"[{current_name}] ❌ Không tìm thấy dòng PROFILE khớp trong Excel – Dừng script.")
                        close_profile(provider, base_url, profile_input.get("id"))
                        return  # DỪNG luôn nếu không có dòng nào trùng

                else:  # mode == 'row'
                    df_valid = df[df.apply(lambda row: any(str(cell).strip() for cell in row), axis=1)]
                    logger(f"[{profile_input.get('name')}] 🧮 Tìm thấy {len(df_valid)} dòng có dữ liệu hợp lệ.")
                    for idx, row in df_valid.iterrows():
                        if stop_flag.is_set():
                            logger(f"[{profile_input.get('name')}] ⛔ Dừng theo yêu cầu trong Excel block.")
                            return

                        row_vars = row.to_dict()
                        row_vars["row_data"] = row_vars.copy()

                        for inner in inner_blocks:
                            if stop_flag.is_set():
                                logger(f"[{profile_input.get('name')}] ⛔ Dừng giữa block Excel.")
                                return
                            execute_block(inner, row_vars)

            except Exception as e:
                logger(f"[EXCEL BLOCK] ❌ Lỗi đọc Excel: {e}")
            return
        
        # end excel

        # save excel

        elif action == 'save_to_excel':
            excel_path = block.get('path')  # Đường dẫn đến file Excel
            profile_column = block.get('profile_column', 'PROFILE')  # Cột PROFILE để tìm dòng
            column_save = block.get('column_save', 'STATUS')  # Cột STATUS để lưu
            value = block.get('value', '')  # Giá trị cần lưu vào cột STATUS
            mode = block.get('mode', 'row')  # Chế độ 'profile' hoặc 'row'
            
            if not excel_path:
                logger(f"[SAVE TO EXCEL] ❌ Không có đường dẫn đến Excel.")
                return

            try:
                # Load workbook và lấy sheet
                wb = openpyxl.load_workbook(excel_path)
                ws = wb.active

                # Tìm cột PROFILE và cột STATUS
                profile_col_idx = None
                status_col_idx = None

                for header_cell in ws[1]:  # Row 1 chứa header
                    if header_cell.value and header_cell.value.strip().lower() == profile_column.strip().lower():
                        profile_col_idx = header_cell.column
                    if header_cell.value and header_cell.value.strip().lower() == column_save.strip().lower():
                        status_col_idx = header_cell.column

                if profile_col_idx is None:
                    logger(f"[{profile_input['name']}] ❌ Không tìm thấy cột {profile_column} trong Excel.")
                    return

                if status_col_idx is None:
                    logger(f"[{profile_input['name']}] ❌ Không tìm thấy cột {column_save} trong Excel.")
                    return

                # Nếu mode là 'profile', tìm dòng của profile
                current_name = profile_input.get("name", "").strip().lower()
                matched = False

                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):  # Bắt đầu từ row 2 (bỏ qua header)
                    profile_cell = row[profile_col_idx - 1].value  # Lấy giá trị cột PROFILE
                    if profile_cell and profile_cell.strip().lower() == current_name:
                        matched = True
                        # Lưu giá trị vào cột STATUS
                        row[status_col_idx - 1].value = value
                        logger(f"[{profile_input['name']}] 💾 Lưu dữ liệu '{value}' vào cột '{column_save}' tại dòng profile {current_name}")
                        break

                if not matched:
                    logger(f"[{profile_input['name']}] ❌ Không tìm thấy dòng PROFILE khớp trong Excel.")
                    return

                # Lưu workbook vào file
                wb.save(excel_path)
                logger(f"[{profile_input['name']}] ✔️ Đã lưu dữ liệu vào Excel tại {excel_path}")

            except Exception as e:
                logger(f"[SAVE TO EXCEL] ❌ Lỗi khi lưu dữ liệu vào Excel: {e}")
            return
        
        # end excel

        xpath = render(block.get('xpath', ''), local_vars)
        value = render(block.get('value', ''), local_vars)

        logger(f"🧠 Giá trị của biến 'variables': {variables}")

        try:
                if action == 'log':
                    logger(f"📝 {value}")

                elif action == 'await_element':
                    try:
                        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, xpath)))
                        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, xpath)))
                        logger(f"⏳ AWAIT ELEMENT thành công: {xpath}")
                    except Exception as e:
                        logger(f"❌ AWAIT ELEMENT thất bại: {xpath} | Lỗi: {e}")
                        if block.get("stop_on_fail", False):
                            logger(f"🛑 DỪNG SCRIPT do 'await_element' thất bại và 'stop_on_fail: true'")
                            close_profile(provider, base_url, profile['id'])
                            try:
                                driver.quit()
                            except Exception as e:
                                logger(f"❌ Lỗi khi đóng trình duyệt: {e}")
                            exit()

                elif action == 'open_url':
                    driver.get(value)
                    logger(f"→ OPEN URL - {value}")

                elif action == 'navigate_back':
                    driver.back()
                    logger(f"⬅️ Quay lại trang trước")
                
                elif action == 'navigate_forward':
                    driver.forward()
                    logger(f"➡️ Tiến tới trang tiếp theo")
                         
                elif action == 'input_text':
                    elem = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, xpath)))
                    elem.clear()
                    elem.send_keys(value)
                    logger(f"→ NHẬP => {xpath} | VALUE = {value}")

                elif action == 'click':
                    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, xpath))).click()
                    logger(f"→ CLICK => {xpath}")

                elif action == 'click_coords':
                    x, y = map(int, value.strip().split(','))
                    webdriver.ActionChains(driver).move_by_offset(x, y).click().perform()
                    logger(f"🖱️ CLICK theo tọa độ: ({x}, {y})")
                    webdriver.ActionChains(driver).move_by_offset(-x, -y).perform()

                elif action == 'input_press_key':
                    elem = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, xpath)))
                    key = getattr(Keys, value.upper(), None)
                    if key:
                        elem.click()
                        elem.send_keys(key)
                        logger(f"→ PRESS KEY {value.upper()} tại {xpath}")
                    else:
                        logger(f"❌ Phím không hợp lệ: {value}")

                elif action == 'element_exists':
                    elements = driver.find_elements(By.XPATH, xpath)
                    exists = len(elements) > 0
                    logger(f"→ ELEMENT {'TỒN TẠI' if exists else 'KHÔNG TỒN TẠI'} => {xpath}")
                    
                    # Nếu phần tử tồn tại, thực hiện các hành động trong if_true
                    if exists and 'if_true' in block:
                        for inner in block['if_true']:
                            execute_block(inner, local_vars)
                    
                    # Nếu phần tử không tồn tại, thực hiện các hành động trong if_false
                    elif not exists and 'if_false' in block:
                        for inner in block['if_false']:
                            execute_block(inner, local_vars)
                    
                    # Nếu phần tử không tồn tại và stop_on_fail = true, dừng script
                    if not exists and block.get("stop_on_fail", False):
                        logger(f"🛑 DỪNG SCRIPT do element không tồn tại và 'stop_on_fail: true'")
                        
                        # Gọi API để đóng profile
                        close_profile(provider, base_url, profile['id'])  # Đóng profile thông qua API
                        
                        # Đảm bảo đóng trình duyệt và thoát khỏi script
                        try:
                            driver.quit()  # Đóng trình duyệt
                        except Exception as e:
                            logger(f"❌ Lỗi khi đóng trình duyệt: {e}")
                        
                        exit()  # Dừng script hoàn toàn
                
                elif action == 'upload_file':
                    file_input = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, xpath)))
                    file_input.send_keys(value)
                    logger(f"📤 Đã tải lên file: {value}")

                elif action == 'wait':
                    time.sleep(float(value))
                    logger(f"🕒 WAIT {value} giây")

                elif action == 'scroll':
                    if value == "down":
                        driver.execute_script("window.scrollBy(0, 300);")
                        logger(f"↓ Scroll xuống 300px")
                    elif value == "up":
                        driver.execute_script("window.scrollBy(0, -300);")
                        logger(f"↑ Scroll lên 300px")
                    elif value == "to_bottom":
                        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                        logger(f"↓ Scroll xuống cuối trang")
                    elif value == "to_top":
                        driver.execute_script("window.scrollTo(0, 0);")
                        logger(f"↑ Scroll lên đầu trang")
                    elif value == "element":
                        if xpath:
                            elem = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, xpath)))
                            driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", elem)
                            logger(f"🔍 Scroll tới element: {xpath}")
                    elif value == "random":
                        direction = random.choice([-1, 1])
                        pixels = random.randint(100, 600)
                        driver.execute_script(f"window.scrollBy(0, {direction * pixels});")
                        logger(f"🎲 Scroll ngẫu nhiên {('↓' if direction == 1 else '↑')} {pixels}px")
                    else:
                        pixels = int(value)
                        driver.execute_script(f"window.scrollBy(0, {pixels});")
                        logger(f"↕ Scroll theo pixel: {pixels}")

                elif action == 'screenshot':
                    os.makedirs('screenshots', exist_ok=True)
                    path = f'screenshots/{profile['name']}_{int(time.time())}.png'
                    driver.save_screenshot(path)
                    logger(f"📸 Screenshot lưu tại: {path}")
                
                elif action == 'eval_script':
                    try:
                        logger(f"⚙️ Thực thi JS: {value[:100]}{'...' if len(value) > 100 else ''}")
                        result = driver.execute_script(value)
                        
                        if 'store_as' in block:
                            var_name = block['store_as']
                            variables[var_name] = result
                            logger(f"🧠 JS Eval → Lưu '{var_name}' = {result}")
                        else:
                            logger(f"🧠 JS Eval → Kết quả: {result}")
                    
                    except Exception as e:
                        logger(f"❌ eval_script lỗi: {type(e).__name__} - {str(e)}")
                        if "TrustedHTML" in str(e):
                            logger("⚠️ Cảnh báo: Trình duyệt đang chặn innerHTML do chính sách bảo mật. Cần tránh dùng innerHTML!")

                elif action == 'loop':
                    count_str = block.get('count', '1')  # Mặc định là 1 nếu không có count
                    count_str = replace_variables_in_string(count_str, variables)  # Thay thế các biến động trong count
                    logger(f"🧠 Giá trị của count sau khi thay thế: {count_str}")  # Log giá trị count đã thay thế
                    try:
                        count = int(count_str)  # Chuyển đổi thành số nguyên
                    except ValueError:
                        logger(f"❌ Lỗi chuyển đổi count sang int: {count_str}")
                        return

                    start = int(block.get('start', 0))
                    var_name = block.get('variable', 'i')
                    loop_blocks = block.get('do', [])
                    
                    for i in range(start, start + count):  # Chạy vòng lặp từ start đến start + count
                        logger(f"🧠 Vòng lặp {i} - Cập nhật biến '{var_name}' = {i}")  # Log giá trị của biến i trong vòng lặp
                        variables[var_name] = i  # Cập nhật biến vòng lặp
                        loop_vars = variables.copy()  # Thay vì local_vars.copy()
                        loop_flags.append({'break': False, 'continue': False})

                        for lb in loop_blocks:
                            if loop_flags[-1]['break']:
                                break
                            if loop_flags[-1]['continue']:
                                loop_flags[-1]['continue'] = False
                                break
                            execute_block(lb, loop_vars)

                        loop_flags.pop()

                elif action == 'while':
                    condition = block.get('condition')
                    var_name = block.get('variable', 'i')
                    loop_blocks = block.get('do', [])

                    condition = replace_variables_in_string(condition, variables)  # Thay thế các biến động trong điều kiện
                    logger(f"🧠 Điều kiện của while: {condition}")


                    while eval(condition, {}, variables):
                        loop_vars = local_vars.copy()
                        loop_vars.update(variables)
                        loop_flags.append({'break': False, 'continue': False})
                        for lb in loop_blocks:
                            if loop_flags[-1]['break']:
                                break
                            if loop_flags[-1]['continue']:
                                loop_flags[-1]['continue'] = False
                                break
                            # Thay thế các biến trong mỗi block trước khi thực thi
                            lb['xpath'] = replace_variables_in_string(lb.get('xpath', ''), variables)
                            lb['value'] = replace_variables_in_string(lb.get('value', ''), variables)
                            logger(f"🧠 Đã thay thế trong block: {lb}")  # Log các khối lệnh đã thay thế
                            
                            execute_block(lb, loop_vars)
                        loop_flags.pop()

                elif action == 'break_loop':
                    if loop_flags:
                        loop_flags[-1]['break'] = True

                elif action == 'next_loop':
                    if loop_flags:
                        loop_flags[-1]['continue'] = True
      
                elif action == 'stop_script':
                    logger(f"🛑 SCRIPT DỪNG LẠI: {value or block.get('reason', 'Không có lý do cụ thể')}")
                    close_profile(provider, base_url, profile['id'])
                    driver.quit()
                    return

                time.sleep(1)

        except Exception as e:
            logger(f"→ ❌ {action} {xpath} => {e}")

    profile = profile_input
    variables = {}
    loop_flags = []

    try:
        chrome_options = Options()
        if isinstance(debugger_address, str):
            chrome_options.debugger_address = debugger_address
        else:
            raise ValueError(f"⚠️ debugger_address không hợp lệ: {debugger_address}")

        driver_dir = os.path.abspath("chromedriver")
        driver_path = os.path.join(driver_dir, "chromedriver.exe")
        service = Service(driver_path)
        driver = webdriver.Chrome(service=service, options=chrome_options)
    except Exception as e:
        logger(f"❌ Không mở được trình duyệt: {e}")
        return

    for block in blocks:
        if stop_flag.is_set():
            logger(f"⛔ Đã nhấn STOP – dừng script ngay lập tức.")
            if 'id' in profile:
                close_profile(provider, base_url, profile['id'])
            try:
                driver.quit()
            except Exception as e:
                logger(f"⚠️ Lỗi khi đóng trình duyệt: {e}")
            return
        execute_block(block, variables)

